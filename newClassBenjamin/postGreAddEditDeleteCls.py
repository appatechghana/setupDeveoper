import openpyxl  # load excel workbook
import os
import shutil
wb = openpyxl.load_workbook("internship.xlsx")
sheet = wb.active  # set sheet1 as active by assumption


list1 = []  # create array for variables
list2 = []  # create array for data types
list3 = []  # create array for shortcodes
list4 = []  # create array for displayNames
list5 = []  # create array for deleteShortcodes

class postGreAddEditDeleteCls:
    __variables = ""
    __dataTypes = ""
    __shortcodes = ""
    __displayName = ""
    __deleteShortcodes = ""

    def __init__(self, tableName, realName, ownerID, action, start):
        self.__tableName = tableName
        self.__realName = realName
        self.__ownerID = ownerID
        self.__action = action
        self.__start = start

    def __postGreAdd(self):
        pgSql = "CREATE OR REPLACE FUNCTION public.sp_" + self.__tableName + "_add(\n\t\t"
        length = len(list3)
        pgSqlSub = list1[0] + " " + list2[0]

        for l in range(1, length):
            pgSqlSub += ",\n\t\t" + list1[l] + " " + list2[l]

        pgSql += pgSqlSub + ",\n\
            \tp_status integer,\n\
            \tp_userid bigint,\n\
            \tp_actionfileid bigint,\n\
            \tp_approvalstateid bigint,\n\
            \tp_approvedbyid bigint,\n\
            \tp_updatereason text,\n\
            \tp_actionresponseid bigint)\n\
             RETURNS SETOF vws_add AS\n\
            $BODY$\n\
            DECLARE\n\
                \tv_rec vws_add%ROWTYPE;\n\
        	    \tv_audit text;\n\
        	    \tv_approval text;\n" \
                            "BEGIN\n" \
                            "/**Insert Data Into Table**/"

        # INSERT TO
        pgSql += "\nINSERT INTO tb_" + self.__tableName + "("
        pgSqlSub = list1[0].replace("p_", "")

        for l in range(1, length):
            pgSqlSub += ",\n\t\t\t\t\t\t\t" + list1[l].replace("p_", "")

        pgSql += pgSqlSub + \
                 ",\n\t\t\t\t\t\t\tstatus," \
                 "\n\t\t\t\t\t\t\tstamp)"

        # VALUES
        pgSql += "\nVALUES ("

        pgSqlSub = list1[0]
        for l in range(1, length):
            pgSqlSub += ",\n\t\t" + list1[l]

        pgSql += pgSqlSub + \
            ",\n\t\tp_status,\n\
            now());\n\n\
            /**Obtain Return Data**/\n\
            SELECT rid,stp \n\
            INTO v_rec\n\
            FROM vw_" + self.__tableName + \
            "\nWHERE rid IN (SELECT currval('tb_" + self.__tableName + "_recid_seq'));" \
            \
            \
            "\n\n/**Prepare Data for Audit **/\n" \
            "SELECT \'Record ID = \'       ||COALESCE(br.rid::varchar,'')\n"

        pgSqlSub = "||' :: " + list4[0] + " = '      ||COALESCE(br." + list3[0] + "::varchar,'')"
        for l in range(1, length):
            pgSqlSub += "\n||' :: " + list4[l] + " = '      ||COALESCE(br." + list3[l] + "::varchar,'')"

        pgSql += pgSqlSub + "\n||' :: Status = '       ||COALESCE(br.sts::varchar,'')\n" \
                "||' :: Date Stamp = '        ||COALESCE(br.stp::varchar,'')\n" \
                "INTO v_audit\n" \
                "\tFROM vw_" + self.__tableName + " br\n" \
                "\t\tWHERE br.rid=v_rec.rid;\n\n" \
                "/**Prepare Data for Approval**/\n" \
                "SELECT (CASE WHEN LOWER(TRIM(COALESCE(br.rid::VARCHAR,''))) = LOWER(TRIM(COALESCE(v_rec.rid::VARCHAR,'')))\n" \
                "THEN 'DELETE FROM tb_" + self.__tableName + " WHERE recid = '||TRIM(COALESCE(v_rec.rid::VARCHAR,'')) ELSE ''END)\n" \
                "INTO v_approval\n" \
                "FROM vw_" + self.__tableName + " br\n" \
                "WHERE br.rid=v_rec.rid;\n\n" \
                "\t\t/**Record Audit**/\n" \
                "\t\tPERFORM fns_audittrail_add(p_userid,'" + self.__realName + " Add',v_audit);\n\n" \
                \
                "\t\t/**Record Approval**/\n" \
                "\t\tPERFORM fns_approvallist_add(p_actionfileid, v_rec.rid, p_approvalstateid, p_approvedbyid, p_userid, \n" \
                "\t\tp_updatereason, v_audit, v_approval, p_actionresponseid);\n\n" \
                \
                "\t\t/**Return Data**/\n" \
                "\t\tRETURN NEXT v_rec;\n" \
                \
                "END;\n" \
                "$BODY$\n" \
                "\tLANGUAGE plpgsql VOLATILE\n" \
                "\tCOST 100\n" \
                "\tROWS 1000;\n" \
                "ALTER FUNCTION public.sp_" + self.__tableName + "_add("

        pgSqlSub = list2[0]
        for l in range(1, length):
            pgSqlSub += ", " + list2[l]

        pgSql += pgSqlSub + ", integer, bigint, bigint, bigint, bigint, text, bigint)\n" \
                            "OWNER TO " + self.__ownerID + ";"
        print self.__tableName + " Add function created successfully"

        file_path = "sqls/" + self.__tableName +"/"+ self.__tableName + "_add.sql"
        directory = os.path.dirname(file_path)

        if not os.path.exists(directory):
            os.makedirs(directory)
        f = open(file_path, "w")
        f.write(pgSql)

    def __postGreEdit(self):
        pgSql = "CREATE OR REPLACE FUNCTION public.sp_" + self.__tableName + "_edit(\n" \
                "\t\tp_recid bigint,\n\t\t"
        length = len(list3)
        pgSqlSub = list1[0] + " " + list2[0]

        for l in range(1, length):
            pgSqlSub += ",\n\t\t" + list1[l] + " " + list2[l]

        pgSql += pgSqlSub + ",\n" \
            "\tp_status integer,\n" \
            "\tp_stamp timestamp without time zone,\n" \
            "\tp_userid bigint,\n" \
            "\tp_actionfileid bigint,\n" \
            "\tp_approvalstateid bigint,\n" \
            "\tp_approvedbyid bigint,\n" \
            "\tp_updatereason text,\n" \
            "\tp_actionresponseid bigint)\n" \
            " RETURNS SETOF vws_edit AS\n" \
            "$BODY$\n" \
            "DECLARE\n" \
            "\t\tv_rec vws_edit%ROWTYPE;\n" \
            "\t\tv_audit text;\n" \
            "\t\tv_approval text;\n" \
            "BEGIN\n" \
            "\t\tv_audit:='';\n" \
            "\t\tv_approval:='';\n\n" \
            \
            "\t\t/**Prepare Data for Approval**/\n" \
            "\t\tSELECT \n" \
            "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br.rid::VARCHAR,''))) = LOWER(TRIM(COALESCE(p_recid::VARCHAR,'')))\n" \
            "\t\t\t\tTHEN 'UPDATE tb_" + self.__tableName+ " SET ' ELSE ''END)||\n"
        pgSqlSub = "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br." + list3[0] + "::VARCHAR,''))) != LOWER(TRIM(COALESCE(" + list1[0] + "::VARCHAR,''))) \n" \
                "\t\t\t\tTHEN ' " + list1[0].replace("p_","") + " = '''||TRIM(COALESCE(br." + list3[0] + "::VARCHAR,''))||''',' ELSE ''END)||\n"

        for l in range(1, length):
            pgSqlSub += "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br." + list3[l] + "::VARCHAR,''))) != LOWER(TRIM(COALESCE(" + list1[l] + "::VARCHAR,''))) \n" \
                "\t\t\t\tTHEN ' " + list1[l].replace("p_","") + " = '''||TRIM(COALESCE(br." + list3[l] + "::VARCHAR,''))||''',' ELSE ''END)||\n"

        pgSql += pgSqlSub + \
                 "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br.sts::VARCHAR,''))) != LOWER(TRIM(COALESCE(p_status::VARCHAR,'')))\n" \
                 "\t\t\t\tTHEN ' status = '||TRIM(COALESCE(br.sts::VARCHAR,''))||',' ELSE ''END)||\n" \
                 "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br.stp::VARCHAR,''))) != LOWER(TRIM(COALESCE(p_stamp::VARCHAR,'')))\n" \
                 "\t\t\t\tTHEN ' stamp = '''||TRIM(COALESCE(br.stp::VARCHAR,'')) || ''' WHERE recid='||TRIM(COALESCE(br.rid::VARCHAR,'')) ELSE ''END)\n" \
                 "\t\tINTO v_approval\n" \
                 "\t\tFROM vw_" + self.__tableName + " br\n" \
                 "\t\tWHERE br.rid=p_recid;\n\n" \
                \
                 "\t\t/**Prepare Data for Audit **/\n" \
                 "\t\tSELECT "

        pgSqlSub = "(CASE WHEN LOWER(TRIM(COALESCE(br." + list3[0] + "::VARCHAR,''))) != LOWER(TRIM(COALESCE(" + list1[0]\
                   + "::VARCHAR,''))) \n" \
                "\t\t\t\tTHEN ' :: " + list4[0] + " (O) = '||TRIM(COALESCE(br." + list3[0] + "::VARCHAR,''))||', (N) = '||TRIM(COALESCE(" + list1[0] + "::VARCHAR,'')) \n" \
                "\t\t\t\tELSE ''END)||\n"
        for l in range(1, length):
            pgSqlSub += "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br." + list3[l] + "::VARCHAR,''))) != LOWER(TRIM(COALESCE(" + \
                        list1[l] + "::VARCHAR,'')))\n" \
                        "\t\t\t\tTHEN ' :: " + list4[l] + " (O) = '||TRIM(COALESCE(br." + list3[l] + "::VARCHAR,''))||', (N) = '||TRIM(COALESCE(" + list1[l] +\
                        "::VARCHAR,''))\n" \
                        "\t\t\t\tELSE ''END)||\n"

        pgSql += pgSqlSub + \
                 "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br.sts::VARCHAR,''))) != LOWER(TRIM(COALESCE(p_status::VARCHAR,'')))\n" \
                 "\t\t\t\tTHEN ' :: Status (O) = '||TRIM(COALESCE(br.sts::VARCHAR,''))||', (N) = '||TRIM(COALESCE(p_status::VARCHAR,''))\n" \
                 "\t\t\t\tELSE ''END)||\n" \
                 "\t\t\t\t(CASE WHEN LOWER(TRIM(COALESCE(br.stp::VARCHAR,''))) != LOWER(TRIM(COALESCE(p_stamp::VARCHAR,'')))\n" \
                 "\t\t\t\tTHEN ' :: Stamp (O) = '||TRIM(COALESCE(br.stp::VARCHAR,''))||', (N) = '||TRIM(COALESCE(p_stamp::VARCHAR,''))\n" \
                 "\t\t\t\tELSE ''END)\n" \
                 "\t\t\t\tINTO v_audit\n" \
                 "\t\tFROM vw_" + self.__tableName + " br\n" \
                 "\t\tWHERE br.rid=p_recid;\n\n" \
                \
                 "UPDATE tb_" + self.__tableName + "\n"
        pgSqlSub = "\tSET " + list1[0].replace("p_", "") + "=" + list1[0]

        for l in range(1, length):
            pgSqlSub += ",\n" + list1[l].replace("p_", "") + "=" + list1[l]

        pgSql += pgSqlSub + ",\nrequireapproval=1,\n" \
                "status=p_status,\n" \
                "stamp=p_stamp\n" \
                " WHERE recid=p_recid;\n\n" \
                \
                "\t\t/** If there is the need for an audit trail, record it **/\n" \
                "\t\tv_audit:='RecId = '||p_recid||v_audit;\n" \
                "\t\tIF NOT(v_audit='') THEN\n" \
                "\t\t\t\tPERFORM fns_audittrail_add(p_userid,'" + self.__realName + " Edit',v_audit);\n\n" \
                \
                "\t\t\t\t/**approval**/\n" \
                "\t\t\t\tPERFORM fns_approvallist_add(p_actionfileid, p_recid, p_approvalstateid,\n" \
                "\t\t\t\tp_approvedbyid, p_userid, p_updatereason, v_audit, v_approval, \n" \
                "\t\t\t\tp_actionresponseid);\n" \
                "\t\tEND IF;\n\n" \
                \
                "\t\t/**Return Data**/\n" \
                "\t\tSELECT rid,stp\n" \
                "\t\t\tINTO v_rec\n" \
                "\t\t\tFROM vw_" + self.__tableName + "\n" \
                "\t\t WHERE rid = p_recid;\n" \
                "\t\tRETURN NEXT v_rec;\n\n" \
                \
                "END;\n" \
                "$BODY$\n" \
                "\tLANGUAGE plpgsql VOLATILE\n" \
                "\tCOST 100\n" \
                "\tROWS 1000;\n" \
                "ALTER FUNCTION public.sp_" + self.__tableName + "_edit(bigint, "

        pgSqlSub = list2[0]
        for l in range(1, length):
            pgSqlSub += ", " + list2[l]

        pgSql += pgSqlSub + ", integer, timestamp without time zone, bigint, bigint, bigint, bigint, text, bigint)\n" \
                            "  OWNER TO " + self.__ownerID + ";"

        print self.__tableName + " Edit function created successfully"

        file_path = "sqls/" + self.__tableName +"/"+ self.__tableName + "_edit.sql"
        directory = os.path.dirname(file_path)

        if not os.path.exists(directory):
            os.makedirs(directory)
        f = open(file_path, "w")
        f.write(pgSql)

    def __postGreDelete(self):
        length = len(list3)
        pgSql = "CREATE OR REPLACE FUNCTION public.sp_" + self.__tableName + "_delete(\n" \
              "\t\tp_recid bigint,\n" \
              "\t\tp_userid bigint,\n" \
              "\t\tp_actionfileid bigint,\n" \
              "\t\tp_approvalstateid bigint,\n" \
              "\t\tp_approvedbyid bigint,\n" \
              "\t\tp_updatereason text,\n" \
              "\t\tp_actionresponseid bigint)\n" \
              "\tRETURNS void AS\n" \
              "$BODY$\n" \
              "DECLARE \n" \
              "\t\tv_rec tb_" + self.__tableName + "%ROWTYPE;\n" \
              "\t\tv_audit text;\n" \
              "\t\tv_approval text;\n" \
              "BEGIN\n\n" \
             \
              "\t/**Prepare Data for Audit **/\n" \
              "\tSELECT 'RecId = '       ||COALESCE(br.rid::varchar,'')\n"

        pgSqlSub = "\t||' :: " + list4[0] + " = '      ||COALESCE(br." + list3[0] + "::varchar,'')\n"
        for l in range(1, length):
            pgSqlSub += "\t||' :: " + list4[l] + " = '  ||COALESCE(br." + list3[l] + "::varchar,'')\n"

        pgSql += pgSqlSub + \
             "\t||' :: status = '       ||COALESCE(br.sts::varchar,'')\n" \
             "\t||' :: stamp = '        ||COALESCE(br.stp::varchar,'')\n" \
             "\t	INTO v_audit\n" \
             "\tFROM vw_" + self.__tableName + " br\n" \
             "\tWHERE br.rid=p_recid;\n\n" \
             \
             "\t/**Obtain Return Data**/\n" \
             "\tSELECT "

        lenCol5 = len(list5)
        pgSqlSub = list5[0]
        for l in range(1, lenCol5):
            pgSqlSub += "," + list5[l]

        pgSql += pgSqlSub + \
                "\n\tINTO v_rec\n" \
                "\tFROM vw_" + self.__tableName + "\n" \
                "\tWHERE rid=p_recid;\n\n" \
                \
                "\t/**Prepare Data for Approval**/\n" \
                "\tSELECT\n" \
                "\t\t (CASE WHEN LOWER(TRIM(COALESCE(br.rid::VARCHAR,''))) = LOWER(TRIM(COALESCE(p_recid::VARCHAR,''))) \n" \
                "\t\t THEN 'INSERT INTO tb_" + self.__tableName + " (recid, "

        pgSqlSub = list1[0].replace("p_", "")
        for l in range(1, length):
            pgSqlSub += ", " + list1[l].replace("p_", "")

        pgSql += pgSqlSub + \
                ", status, stamp) VALUES('||v_rec.recid||','''||\n" \
                "\t\t v_rec."

        pgSqlSub = list1[0].replace("p_", "") + "||'''"
        for l in range(1, length):
            pgSqlSub += ",'''||v_rec." + list1[l].replace("p_", "") + "||'''"

        pgSql += pgSqlSub + \
                 ",'||v_rec.status||','''||v_rec.stamp||''')' ELSE ''END)\n" \
                 "\tINTO v_approval\n" \
                 "\tFROM vw_" + self.__tableName + " br\n" \
                 "\tWHERE br.rid=p_recid;\n\n" \
            \
                "\t/**Delete Record **/\n" \
                "\tDELETE FROM tb_" + self.__tableName + " WHERE recid=p_recid;\n" \
            \
                "\t/**Record Audit**/\n" \
                "\tIF FOUND THEN\n" \
                "\t	PERFORM fns_audittrail_add(p_userid,'" + self.__realName + " Delete',v_audit);\n\n" \
            \
                 "\t\t/**approval**/\n" \
                 "\t\tPERFORM fns_approvallist_add(p_actionfileid, p_recid, p_approvalstateid,\n" \
                 "\t\tp_approvedbyid, p_userid, p_updatereason, v_audit, v_approval,\n" \
                 "\t\tp_actionresponseid);\n" \
                 "\tEND IF;\n\n" \
            \
                 "\tRETURN;\n" \
                 "END;\n" \
                 "$BODY$\n" \
                 "  LANGUAGE plpgsql VOLATILE\n" \
                 "  COST 100;\n" \
                 "ALTER FUNCTION public.sp_" + self.__tableName + "_delete(bigint, bigint, bigint, bigint, bigint, text, bigint)\n" \
                 "  OWNER TO " + self.__ownerID + ";"

        print self.__tableName + " Delete function created successfully"
        file_path = "sqls/" + self.__tableName + "/"+ self.__tableName +"_delete.sql"
        directory = os.path.dirname(file_path)

        if not os.path.exists(directory):
            os.makedirs(directory)
        f = open(file_path, "w")
        f.write(pgSql)

    def postGreAddEditDelete(self):
        global list1
        global list2
        global list3
        global list4
        global list5

        if (self.__action == "a"):
            i =self.__start

            while (sheet.cell(column=1, row=i).value <> None):
                __variables = sheet.cell(column=1, row=i).value
                __dataTypes = sheet.cell(column=2, row=i).value
                __shortcodes = sheet.cell(column=3, row=i).value
                __displayName = sheet.cell(column=4, row=i).value
                list1.append(str(__variables))
                list2.append(str(__dataTypes))
                list3.append(str(__shortcodes))
                list4.append(str(__displayName))
                i += 1
            postGreAddEditDeleteCls.__postGreAdd(self)

        elif (self.__action == "e"):
            i = self.__start

            while (sheet.cell(column=1, row=i).value <> None):
                __variables = sheet.cell(column=1, row=i).value
                __dataTypes = sheet.cell(column=2, row=i).value
                __shortcodes = sheet.cell(column=3, row=i).value
                __displayName = sheet.cell(column=4, row=i).value

                list1.append(str(__variables))
                list2.append(str(__dataTypes))
                list3.append(str(__shortcodes))
                list4.append(str(__displayName))
                i += 1

            postGreAddEditDeleteCls.__postGreEdit(self)


        elif (self.__action == "d"):
            i = self.__start

            while (sheet.cell(column=1, row=i).value <> None):
                __variables = sheet.cell(column=1, row=i).value
                __dataTypes = sheet.cell(column=2, row=i).value
                __shortcodes = sheet.cell(column=3, row=i).value
                __displayName = sheet.cell(column=4, row=i).value

                list1.append(str(__variables))
                list2.append(str(__dataTypes))
                list3.append(str(__shortcodes))
                list4.append(str(__displayName))
                i += 1
            j = self.__start
            while (sheet.cell(column=5, row=j).value <> None):
                __deleteShortcodes = sheet.cell(column=5, row=j).value
                list5.append(str(__deleteShortcodes))
                j += 1
            postGreAddEditDeleteCls.__postGreDelete(self)
        list1 = []
        list2 = []
        list3 = []
        list4 = []
        list5 = []


def pgFunc(tableName, realName, ownerID, start,add = "",edit = "",delete = ""):
    __pgList = [add, edit, delete]
    length = len(__pgList)

    for i in range(0, length):
        if (__pgList[i] == "a"):
            pgAdd = postGreAddEditDeleteCls(tableName, realName, ownerID, "a", start)
            pgAdd.postGreAddEditDelete()

        elif (__pgList[i] == "e"):
            pgEdt = postGreAddEditDeleteCls(tableName, realName, ownerID, "e", start)
            pgEdt.postGreAddEditDelete()

        elif (__pgList[i] == "d"):
            pgDel = postGreAddEditDeleteCls(tableName, realName, ownerID, "d", start)
            pgDel.postGreAddEditDelete()

def identifiers(start,add = "",edit = "", delete = ""):
    i = start
    list6 = []

    while (sheet.cell(column=5, row=i).value <> None):
        tableName = sheet.cell(column=6, row=i).value

        if (sheet.cell(column=6, row=i).value <> None):
            list6.append(str(tableName))
        i += 1

    file_path = "sqls/" + list6[0] + "/" + list6[0] + "_delete.sql"
    directory = os.path.dirname(file_path)

    if os.path.exists(directory):
        shutil.rmtree(directory)
        os.makedirs(directory)

    pgFunc(list6[0], list6[1], "investment",start, add, edit, delete)
    return i + 1
